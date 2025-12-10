import requests
import json
import openpyxl
from datetime import datetime, timedelta, time

with open('cookies.json', 'r') as f:
    cookies_list = json.load(f)
cookies = {cookie['name']: cookie['value'] for cookie in cookies_list}

# Common Headers
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36',
    'Accept': '*/*',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'X-Requested-With': 'XMLHttpRequest',
    'Connection': 'keep-alive',
    'Referer': 'https://activity-enrichment.apps.binus.ac.id/LearningPlan/StudentIndex',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'Pragma': 'no-cache',
    'Cache-Control': 'no-cache',
    'sec-ch-ua': '"Google Chrome";v="143", "Chromium";v="143", "Not A(Brand";v="24"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"'
}

def get_months():
    url = 'https://activity-enrichment.apps.binus.ac.id/LogBook/GetMonths'
    response = requests.get(url, headers=headers, cookies=cookies)
    print("get_months status:", response.status_code)
    if response.status_code == 200:
        data = response.json()['data']
        print("Months fetched:", [m['month'] for m in data])
        return data
    else:
        print(f"Failed to get months: {response.status_code}")
        return []

def get_logbook(header_id):
    url = 'https://activity-enrichment.apps.binus.ac.id/LogBook/GetLogBook'
    data = f'logBookHeaderID={header_id}'
    post_headers = headers.copy()
    post_headers['Content-Type'] = 'application/x-www-form-urlencoded; charset=UTF-8'
    post_headers['Origin'] = 'https://activity-enrichment.apps.binus.ac.id'
    response = requests.post(url, headers=post_headers, data=data, cookies=cookies)
    print("get_logbook status for", header_id, ":", response.status_code)
    if response.status_code == 200:
        data = response.json()['data']
        print("Logbook entries:", len(data))
        return data
    else:
        print(f"Failed to get logbook for {header_id}: {response.status_code}")
        return []

def time_to_decimal(t):
    if isinstance(t, time):
        return (t.hour + t.minute / 60 + t.second / 3600) / 24
    return t

def decimal_to_time(dec):
    if dec == 'OFF' or dec is None:
        return 'OFF'
    total_minutes = int(float(dec) * 1440)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    if hours < 12:
        period = 'AM'
        if hours == 0:
            hours = 12
    else:
        period = 'PM'
        hours -= 12
        if hours == 0:
            hours = 12
    return f"{hours:02d}:{minutes:02d} {period}"

def save_entry(header_id, date_str, clock_in, clock_out, activity, description, entry_id):
    url = 'https://activity-enrichment.apps.binus.ac.id/LogBook/StudentSave'
    date_iso = f"{date_str}T00:00:00"
    post_data = {
        "model[ID]": entry_id,
        "model[LogBookHeaderID]": header_id,
        "model[Date]": date_iso,
        "model[Activity]": activity,
        "model[ClockIn]": clock_in,
        "model[ClockOut]": clock_out,
        "model[Description]": description,
        "model[flagjulyactive]": "false"
    }
    post_headers = headers.copy()
    post_headers['Content-Type'] = 'application/x-www-form-urlencoded; charset=UTF-8'
    post_headers['Origin'] = 'https://activity-enrichment.apps.binus.ac.id'
    response = requests.post(url, headers=post_headers, data=post_data, cookies=cookies)
    print("save_entry status for", date_str, ":", response.status_code, response.text)
    if response.status_code == 200:
        result = response.json()
        print(f"Save response for {date_str}: {result['status']}")
    else:
        print(f"Failed to save entry for {date_str}: {response.status_code}")

months = get_months()
month_to_header = {}
for m in months:
    month_name = m['month']
    header_id = m['logBookHeaderID']
    month_to_header[month_name] = header_id
print("Month to header:", month_to_header)

wb = openpyxl.load_workbook('monthly_activity.xlsx', data_only=True)
ws = wb['Sheet1']
print("Excel loaded, rows:", ws.max_row)

for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    print(f"Row {row_num}: {row}")
    serial, activity, description, clock_in_dec, clock_out_dec = row + (None,) * (5 - len(row))
    
    clock_in_dec = time_to_decimal(clock_in_dec)
    clock_out_dec = time_to_decimal(clock_out_dec)
    
    if serial is None:
        print("Skipping row due to empty serial")
        continue
    
    if isinstance(serial, datetime):
        date_obj = serial
    else:
        try:
            serial = float(serial)
            date_obj = datetime(1899, 12, 30) + timedelta(days=serial)
        except (ValueError, TypeError):
            print("Skipping row due to invalid serial")
            continue
    
    date_str = date_obj.strftime('%Y-%m-%d')
    print("Processing date:", date_str)
    dt = datetime.strptime(date_str, '%Y-%m-%d')
    month_name = dt.strftime('%B')
    header_id = month_to_header.get(month_name)
    if not header_id:
        print(f"No header for month {month_name}")
        continue
    
    clock_in = decimal_to_time(clock_in_dec)
    clock_out = decimal_to_time(clock_out_dec)
    print("Times:", clock_in, clock_out)
    
    logbook = get_logbook(header_id)
    entry_id = '00000000-0000-0000-0000-000000000000'
    found_entry = False
    for entry in logbook:
        entry_date = entry['date'][:10]
        if entry_date == date_str:
            found_entry = True
            entry_id = entry['id']
            break
    
    if found_entry:
        print("Updating entry for", date_str)
    else:
        print("Creating new entry for", date_str)
    
    save_entry(header_id, date_str, clock_in, clock_out, activity, description, entry_id)